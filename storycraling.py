import os
import time
import markdownify
from urllib.parse import urlparse

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 🔹 엑셀 파일 이름
excel_path = r"C:\Users\leehyein\Documents\code\kia\AI-Response-Enhancement\hyundea.story.xlsx"

# 🔹 저장할 마크다운 파일 경로
output_md_file = "kia_all_articles.md"

# 🔹 파일명 정리 함수 (본문 안에 참고용으로만 사용)
def get_url_path_id(url):
    parsed = urlparse(url)
    return parsed.path.lstrip("/").replace("/", " · ")

# 🔹 셀레니움 설정
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# 🔹 엑셀에서 URL 읽기
wb = load_workbook(excel_path)
ws = wb.active
urls = [cell.value for cell in ws['A'] if cell.value and str(cell.value).startswith("http")]

# 🔹 결과 저장할 문자열 리스트
all_markdown = []

for base_url in urls:
    try:
        driver.get(base_url)
        time.sleep(2)
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, "html.parser")

        # 불필요한 태그 제거
        for tag in soup.find_all(["header", "footer", "img", "style", "aside", "nav", "from"]):
            tag.decompose()

        md_content = markdownify.markdownify(str(soup), heading_style="ATX")

        markdown_block = f"""url: {base_url}  
id: {get_url_path_id(base_url)}  
text:  
{md_content}

---
"""
        all_markdown.append(markdown_block)
        print(f"✅ 크롤링 완료: {base_url}")

    except Exception as e:
        print(f"❌ 실패: {base_url} | 이유: {e}")

driver.quit()

# 🔹 마크다운 파일 하나로 저장
with open(output_md_file, "w", encoding="utf-8") as f:
    f.writelines(all_markdown)

print(f"\n📄 전체 결과 저장 완료 → {output_md_file}")
