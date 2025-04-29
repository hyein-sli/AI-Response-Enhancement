import os
import time
import markdownify
from urllib.parse import urlparse

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 🔹 엑셀 파일 경로
excel_path = r"C:\Users\leehyein\Documents\code\kia\AI-Response-Enhancement\hyundae.news.xlsx"

# 🔹 저장 폴더
output_folder = "kia_articles_md"
os.makedirs(output_folder, exist_ok=True)

# 🔹 파일 이름 추출 함수
def get_news_id(url):
    parsed = urlparse(url)
    parts = parsed.path.strip("/").split("/")
    for part in parts:
        if part.startswith("CONT"):
            return f"news.{part}.md"
    return "news.unknown.md"

# 🔹 셀레니움 설정
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# 🔹 엑셀에서 URL 읽기
wb = load_workbook(excel_path)
ws = wb.active
urls = [cell.value for cell in ws['A'] if cell.value and str(cell.value).startswith("http")]

# 🔹 본문 크롤링 및 마크다운 저장
for base_url in urls:
    try:
        driver.get(base_url)
        time.sleep(2)
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, "html.parser")

        # 기사 제목 추출
        title_tag = soup.find("h2", class_="news-detail__title")
        title = title_tag.get_text(strip=True) if title_tag else "(제목 없음)"

        # 불필요한 태그 제거
        for tag in soup.find_all(["header", "footer", "img", "style", "aside", "nav", "from"]):
            tag.decompose()

        # HTML → 마크다운 변환
        md_body = markdownify.markdownify(str(soup), heading_style="ATX")

        # 마크다운 구성
        markdown_text = f"""# {base_url}

## {title}

{md_body.strip()}
"""

        # 파일 저장
        file_name = get_news_id(base_url)
        file_path = os.path.join(output_folder, file_name)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(markdown_text)

        print(f"✅ 저장 완료: {file_name}")

    except Exception as e:
        print(f"❌ 실패: {base_url} | 이유: {e}")

driver.quit()
print(f"\n📁 모든 마크다운 파일이 '{output_folder}' 폴더에 저장되었습니다.")
