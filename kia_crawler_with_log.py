import os
import time
import markdownify
from urllib.parse import urlparse

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 📁 저장 폴더 (없으면 생성)
md_folder = './members_kia'
if not os.path.exists(md_folder):
    os.makedirs(md_folder)

# 🔹 파일명 생성 함수
def sanitize_filename(url):
    parsed = urlparse(url)
    path = parsed.path
    if not path or path == "/":
        filename = "index.md"
    else:
        filename = path.lstrip("/").replace("/", ".")
        if not filename.endswith(".md"):
            filename += ".md"
    return filename

# 🔹 셀레니움 설정
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# 🔹 엑셀에서 URL 리스트 불러오기
excel_path = r"C:\Users\혜민\Documents\code\kia\AI-Response-Enhancement/members.kia.xlsx"
wb = load_workbook(excel_path)
ws = wb.active
urls = [cell.value for cell in ws['A'] if cell.value and str(cell.value).startswith("http")]

# 🔁 각 URL 순회
for url in urls:
    try:
        driver.get(url)
        time.sleep(2)  # 로딩 대기

        html_content = driver.page_source
        soup = BeautifulSoup(html_content, "html.parser")

        # 불필요한 태그 제거
        for tag in soup.find_all(["header", "footer", "img", "style", "aside", "nav", "from"]):
            tag.decompose()

        md_content = markdownify.markdownify(str(soup), heading_style="ATX")
        filename = sanitize_filename(url)
        filepath = os.path.join(md_folder, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(md_content)

        print(f"✅ Saved: {url} → {filepath}")

    except Exception as e:
        print(f"❌ Error fetching {url}: {e}")

driver.quit()
